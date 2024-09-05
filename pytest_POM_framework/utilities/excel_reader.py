import pytest
import openpyxl


# usernames=["reddyvinuth27@gmail.com","abc@gmail.com","xyz@gmail.com","mno@gmail.com"]
# usernames=[("reddyvinuth27@gmail.com","selenium"),("abc@gmail.com","1234"),("xyz@gmail.com","1123"),("mno@gmail.com","1234567")]

def get_list_from_excel(excel_name,sheet_name):
    excel = openpyxl.load_workbook(excel_name)
    sheet = excel[sheet_name]
    # login_creds.cell(1,2)
    values = []
    for row in range(1, sheet.max_row+1):
        nested_values = []
        for column in range(1, sheet.max_column+1):
            data = sheet.cell(row, column)
            nested_values.append(data.value)
        values.append(nested_values)
    return values

def add_value_to_excel(excel_name,sheet_name):
    excel = openpyxl.load_workbook(excel_name)
    sheet = excel[sheet_name]
    sheet.cell(7, 1).value = "new value"
    excel.save(excel_name)
add_value_to_excel("creds.xlsx","login_creds")